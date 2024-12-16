import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from tkcalendar import DateEntry
import babel.numbers

class LogParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CTR Trend Chart Generator v1.0.1")

        # 閥門編號
        self.valve_ids = ["P1-1", "P2-1", "P3-1", "P4-1", "P9-1", "P9-2", "P9-3", "P10-1", "P10-2", "P10-3"]

        # 預設解析格式為 MSC 2.x
        self.selected_format = tk.StringVar(value="MSC 2.x")
        
        # 設定起始與結束日期變數
        self.start_date = None
        self.end_date = None
        self.log_data = None  # 用於存儲所有解析的數據
        self.min_time = None
        self.max_time = None

        # 設定GUI元件
        self.setup_gui()

    def setup_gui(self):
        # 建立一個容器 Frame，將按鈕水平排列並置中
        self.button_container = tk.Frame(self.root)
        self.button_container.pack(pady=5)

        # 解析格式按鈕框架
        self.format_frame = tk.LabelFrame(self.button_container, text="Log Format")
        self.format_frame.pack(side=tk.LEFT, padx=5)  # 左邊，增加間距

        # 選擇資料夾按鈕
        self.select_folder_button = tk.Button(
            self.button_container, text="Select Log Folder", command=self.select_log_folder
        )
        self.select_folder_button.pack(side=tk.LEFT, padx=5)  # 靠近解析格式框架
        
        tk.Radiobutton(self.format_frame, text="MSC 2.x", variable=self.selected_format, value="MSC 2.x").pack(anchor="w")
        tk.Radiobutton(self.format_frame, text="MSC 3.x", variable=self.selected_format, value="MSC 3.x").pack(anchor="w")

        # 進度條
        self.progress_bar = ttk.Progressbar(self.root, orient="horizontal", mode="determinate", length=300)
        self.progress_bar.pack(pady=5)

        # 時間區間顯示與選擇
        self.date_selection_frame = tk.Frame(self.root)
        self.date_selection_frame.pack(pady=5)

        tk.Label(self.date_selection_frame, text="Start Date").grid(row=0, column=0, padx=5)
        tk.Label(self.date_selection_frame, text="End Date").grid(row=1, column=0, padx=5)

        self.start_date_entry = DateEntry(self.date_selection_frame, date_pattern="yyyy/mm/dd")
        self.start_date_entry.grid(row=0, column=1, padx=5)
        self.start_date_entry.bind("<<DateEntrySelected>>", lambda e: self.validate_date_range(self.start_date_entry, self.min_time.date(), self.max_time.date()))
        
        self.start_hour_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=23, width=3, format="%02.0f")
        self.start_hour_spinbox.grid(row=0, column=2)
        tk.Label(self.date_selection_frame, text="Hr").grid(row=0, column=3, padx=2)

        self.start_minute_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=59, width=3, format="%02.0f")
        self.start_minute_spinbox.grid(row=0, column=4)
        tk.Label(self.date_selection_frame, text="Min").grid(row=0, column=5, padx=2)

        self.start_second_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=59, width=3, format="%02.0f")
        self.start_second_spinbox.grid(row=0, column=6)
        tk.Label(self.date_selection_frame, text="Sec").grid(row=0, column=7, padx=2)

        self.end_date_entry = DateEntry(self.date_selection_frame, date_pattern="yyyy/mm/dd")
        self.end_date_entry.grid(row=1, column=1, padx=5)
        self.end_date_entry.bind("<<DateEntrySelected>>", lambda e: self.validate_date_range(self.end_date_entry, self.min_time.date(), self.max_time.date()))

        self.end_hour_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=23, width=3, format="%02.0f")
        self.end_hour_spinbox.grid(row=1, column=2)
        tk.Label(self.date_selection_frame, text="Hr").grid(row=1, column=3, padx=2)

        self.end_minute_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=59, width=3, format="%02.0f")
        self.end_minute_spinbox.grid(row=1, column=4)
        tk.Label(self.date_selection_frame, text="Min").grid(row=1, column=5, padx=2)

        self.end_second_spinbox = tk.Spinbox(self.date_selection_frame, from_=0, to=59, width=3, format="%02.0f")
        self.end_second_spinbox.grid(row=1, column=6)
        tk.Label(self.date_selection_frame, text="Sec").grid(row=1, column=7, padx=2)

        # 匯出Excel按鈕
        self.export_button = tk.Button(self.root, text="Export to Excel", command=self.export_to_excel, state="disabled")
        self.export_button.pack(pady=10)

    def validate_date_range(self, entry, mindate, maxdate):
        selected_date = entry.get_date()
        if selected_date < mindate:
            messagebox.showwarning("Warning", f"Selected date is before the start of available range ({mindate}). Resetting.")
            entry.set_date(mindate)
        elif selected_date > maxdate:
            messagebox.showwarning("Warning", f"Selected date is after the end of available range ({maxdate}). Resetting.")
            entry.set_date(maxdate)

    def select_log_folder(self):
        # 選擇日誌檔案資料夾
        folder_path = filedialog.askdirectory(title="Select Log Files Folder")
        if not folder_path:
            return

        # 解析全部日誌文件並顯示進度
        self.progress_bar["value"] = 0
        self.root.update_idletasks()
        
        self.log_data, time_range = self.parse_log_files(folder_path)
        if not self.log_data:
            messagebox.showerror("Error", "No valid log files found in the selected folder.")
            return

        # 獲取時間範圍
        self.min_time = min(time_range)
        self.max_time = max(time_range)
        self.start_date_entry.config(mindate=self.min_time.date(), maxdate=self.max_time.date())
        self.end_date_entry.config(mindate=self.min_time.date(), maxdate=self.max_time.date())

        # 提示用戶
        messagebox.showinfo("Info", f"Log files successfully parsed. \nAvailable time range: {self.min_time} to {self.max_time}.")
        self.export_button.config(state="normal")

    def parse_log_files(self, folder_path):
        # 選擇解析格式
        selected_format = self.selected_format.get()
        if selected_format == "MSC 2.x":
            regex_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \(\d+\) PressEvTh\(\): Sent MULTIJET_EVENT_CODE_CURRENT_PRESSURE_HAS_CHANGED\((\d+), (\d+), press=(-?\d+\.\d+), array=(-?\d+\.\d+)\)")
        elif selected_format == "MSC 3.x":
            regex_pattern = re.compile(r"(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}:\d{2}\.\d+): \(.*?\) MultiJetImpl::MCPressCurrentValueChangedEvent\((\d+),(\d+)\), .*?pressure = (-?\d+\.\d+) mbar.*")
        else:
            raise ValueError("Unsupported log format selected.")
        
        log_data = {valve_id: [] for valve_id in self.valve_ids}
        
        # 用於存儲檔案中的時間戳
        time_stamps = []
        log_files = [f for f in os.listdir(folder_path) if re.match(r"^mjnxtdebug\d{8}\.log$", f)]
        total_files = len(log_files)

        # 遍歷資料夾中的日誌檔案
        for idx, file_name in enumerate(log_files, 1):
            file_path = os.path.join(folder_path, file_name)

            with open(file_path, 'r') as file:
                for line in file:
                    match = regex_pattern.search(line)
                    if match:
                        date_time_str, main_id, sub_id, press_value = match.groups()[:4]
                        parsed_time = datetime.strptime(date_time_str.split('.')[0], "%Y/%m/%d, %H:%M:%S")
                        time_stamps.append(parsed_time)
                        
                        valve_id = f"P{main_id}-{sub_id}"
                        if valve_id in log_data:
                            log_data[valve_id].append((parsed_time, float(press_value)))

            # 更新進度條
            self.progress_bar["value"] = (idx / total_files) * 100
            self.root.update_idletasks()

        # 檢查是否有有效的數據
        if not time_stamps:
            return None, None

        return log_data, sorted(set(time_stamps))  # 確保時間點是有序且唯一的

    def export_to_excel(self):
        # 選擇開始和結束時間
        start_date = self.start_date_entry.get()
        start_time = f"{self.start_hour_spinbox.get()}:{self.start_minute_spinbox.get()}:{self.start_second_spinbox.get()}"
        end_date = self.end_date_entry.get()
        end_time = f"{self.end_hour_spinbox.get()}:{self.end_minute_spinbox.get()}:{self.end_second_spinbox.get()}"

        start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y/%m/%d %H:%M:%S")
        end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y/%m/%d %H:%M:%S")

        # 過濾數據並選擇輸出路徑
        filtered_data = self.filter_data(start_datetime, end_datetime)
        if not filtered_data:
            messagebox.showwarning("Warning", "No data in the selected time range.")
            return
        
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            return
        
        self.save_to_excel(filtered_data, save_path)

        messagebox.showinfo("Success", "Data has been successfully exported to Excel.")

    def filter_data(self, start_datetime, end_datetime):
        filtered_data = {}
        for valve_id, data in self.log_data.items():
            filtered_data[valve_id] = [(dt, val) for dt, val in data if start_datetime <= dt <= end_datetime]
        return filtered_data

    def save_to_excel(self, filtered_data, save_path):
        # 建立新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "All Valve Data"  # 設定工作表名稱

        # 設定表頭
        ws.append(["Date Time"] + [f"{valve_id} Press Value" for valve_id in filtered_data.keys()])  # 依序為每個閥門添加壓力數據的標題

        # 收集所有時間點
        all_times = sorted(set(time for data in filtered_data.values() for time, _ in data))

        total_steps = len(all_times) + 1  # 時間點數量 + 一個繪製圖表步驟
        self.progress_bar["maximum"] = total_steps  # 設定進度條的最大值
        self.progress_bar["value"] = 0

        # 填充數據
        for row_idx, time in enumerate(all_times, start=2):
            ws.cell(row=row_idx, column=1, value=time)  # 寫入時間
            for col_idx, valve_id in enumerate(filtered_data.keys(), start=2):
                # 嘗試找到該時間點對應的壓力值
                press_value = next((val for dt, val in filtered_data[valve_id] if dt == time), None)
                if press_value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=press_value)
                else:
                    # 若無對應壓力值，計算前後壓力值的平均值
                    earlier_values = [val for dt, val in filtered_data[valve_id] if dt < time]
                    later_values = [val for dt, val in filtered_data[valve_id] if dt > time]

                    # 獲取最近的前後值
                    earlier_value = earlier_values[-1] if earlier_values else None
                    later_value = later_values[0] if later_values else None

                    # 計算平均值並填入
                    if earlier_value is not None and later_value is not None:
                        average_value = (earlier_value + later_value) / 2
                        ws.cell(row=row_idx, column=col_idx, value=average_value)
                    elif earlier_value is not None:  # 若只有前值
                        ws.cell(row=row_idx, column=col_idx, value=earlier_value)
                    elif later_value is not None:  # 若只有後值
                        ws.cell(row=row_idx, column=col_idx, value=later_value)

            # 更新進度條
            self.progress_bar["value"] += 1
            self.root.update_idletasks()
            #self.progress_bar["maximum"] = 0

        # 建立趨勢圖
        chart = LineChart()
        chart.title = "Valve Pressure Trends"
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Press Value (mbar)"

        # 設定數據範圍，從第二列開始，以包含所有閥門的數據
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(filtered_data), max_row=len(all_times) + 1)
        chart.add_data(data_ref, titles_from_data=True)  # 包含標題

        # 設定 X 軸標籤（時間範圍）
        time_ref = Reference(ws, min_col=1, min_row=2, max_row=len(all_times) + 1)
        chart.set_categories(time_ref)

        # 設定 X 軸標籤格式
        chart.x_axis.number_format = "yyyy/mm/dd hh:mm:ss"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickLblPos = "low"

        # 添加趨勢圖到工作表
        ws.add_chart(chart, "N2")  # 設定圖表顯示位置

        # 更新進度條（完成繪製圖表步驟）
        self.progress_bar["value"] += 1
        self.root.update_idletasks()

        # 儲存 Excel 檔案
        wb.save(save_path)

        # 完成處理後，重置進度條
        self.progress_bar["maximum"] = 0

if __name__ == "__main__":
    root = tk.Tk()
    app = LogParserApp(root)
    root.mainloop()